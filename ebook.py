#!/bin/python
#-*- coding: UTF-8 -*-
import re
import random
import traceback
import os 
import codecs
import sys
import time
import urllib
import urllib.parse
import urllib.error
import urllib.request
import chardet
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
import importlib
importlib.reload(sys)


#Some User Agents
hds=[{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},\
{'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},\
{'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]


'''
	接口用于返回page页面的所有urlId
	idList:	保存id列表
	page：	起始页面入参
	Author:	icoty/yangyu    http://www.2belief.com/
	https://www.77169.com/html/170460.html
	http://www.cnblogs.com/zhaof/p/7326260.html
	http://litten.me/2017/07/09/prevent-spiders/
'''
def getDownloadID(idList, page):
    id = 0
    try_times=0

    index = []
    i = page
    while i < page+1:
        index.append(i)
        i += 1
    random.shuffle(index)

    while(1):
        url='http://mebook.cc/page/'+str(index[id])
        time.sleep(np.random.rand()*5)

        try:
            req = urllib.request.Request(url, headers=hds[id%len(hds)])
            source_code = urllib.request.urlopen(req).read().decode('utf-8')
            plain_text=str(source_code)   
        except urllib.error.HTTPError as e:
            print(e.code)
            print(e.reason)
            continue
        except urllib.error.URLError as e:
            print(e.code)
            print(e.reason)
            continue

        soup = BeautifulSoup(plain_text,'html.parser')
        list_soup = soup.find_all('div', {'class': 'thumbnail'})	# 一个page内的全部书籍

        try_times += 1
        if list_soup == None and try_times < 5:
            continue
        elif list_soup == None or len(list_soup) < 1:
            break

        for book_info in list_soup:
            book_url = book_info.find('div', {'class': 'img'}).find('a').get('href')
            down_id = re.findall('\d+',book_url)
            idList.append(eval(down_id[0]))
            try_times = 0
        print('get page number %d ' % index[id])
        id += 1
        if(id >= len(index)):break
    idList = list(set(idList))


'''
	接口用于返回page页面的所有urlId
	idList:	保存id列表
	page：	起始页面
	Author:	icoty/yangyu    http://www.2belief.com/
'''
def spider(idList, page):
    wb = load_workbook("ebook-url.xlsx")
    ws = wb.get_sheet_by_name('ebook')
    ws_rows_len = ws.max_row

    id = 0
    try_times=0
    random.shuffle(idList)
    print('next download count %d' % len(idList))

    while(1):
        url='http://mebook.cc/download.php?id='+str(idList[id])
        time.sleep(np.random.rand()*5)

        try:
            req = urllib.request.Request(url, headers=hds[id%len(hds)])
            source_code = urllib.request.urlopen(req).read().decode('utf-8')
            plain_text=str(source_code)   
        except urllib.error.HTTPError as e:
            print(e.code)
            print(e.reason)
            continue
        except urllib.error.URLError as e:
            print(e.code)
            print(e.reason)
            continue

        soup = BeautifulSoup(plain_text,'html.parser')
        #with codecs.open('pan.html', 'w+', 'utf-8') as f:
        #    f.write(str(soup))

        # 形如 <title>《苏轼研究：王水照苏轼研究四种》王水照（作者）epub+mobi+azw3下载页面</title>
        book_name = soup.head.title
        if None != book_name:
            book_name = book_name.string.strip().split('epub')[0]

        # 形如 <p>网盘密码：百度网盘密码：pj14     天翼云盘密码：5134</p>
        url_pwd = soup.find('div', {'class': 'desc'}).find_all('p')[-2].string.split('：')
        tiany_pwd = ''
        baidu_pwd = ''
        if (4 == len(url_pwd)):
            baidu_pwd = ' (访问码:' + url_pwd[2].split()[0] + ')'
            tiany_pwd = ' (访问码:' + url_pwd[-1].split()[0]+ ')'
        elif (3 == len(url_pwd)):
            baidu_pwd = ' (访问码:' + url_pwd[2].split()[0] + ')'

        try_times += 1
        list_soup = soup.find('div', {'class': 'list'}).find_all('a')	# 各盘的下载链接列表
        if None==list_soup and try_times<5:
            continue
        elif None==list_soup or len(list_soup)<1:
            with codecs.open('error.txt', 'a+', 'utf-8') as f:
                f.write(str(list_soup))
                f.write('\n')
            print('list_soup：%s' % str(list_soup))
            id += 1
            if id >= len(idList):break
            continue

        tiany_url = ''
        micro_url = ''
        baidu_url = str(list_soup[0]).split('\"')[1] + baidu_pwd
        if (2 == len(list_soup)):
            with codecs.open('error.txt', 'a+', 'utf-8') as f:
                f.write(str(list_soup))
                f.write('\n')
            micro_url = str(list_soup[1]).split('\"')[1]
            try_times=0
        elif (3 == len(list_soup)):
            micro_url = str(list_soup[1]).split('\"')[1]
            tiany_url = str(list_soup[2]).split('\"')[1] + tiany_pwd
            try_times=0

        ws.append([int(ws_rows_len),book_name,baidu_url])
        ws_rows_len += 1

        print('get download id %d ' % idList[id])
        id += 1
        if id >= len(idList):break
    save_path = 'ebook-url.xlsx'
    wb.save(save_path)


if __name__=='__main__':
    i = 0
    while i < 2:	# 代理不稳定容易断开,采用小批量多批次,防止被禁,如果你有私密代理可以自行设置代理大批量爬取
        idList = []
        getDownloadID(idList, i*10)
        spider(idList, i*10) 
        i += 1